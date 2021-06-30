import cv2
import os
import numpy as np
from openpyxl import load_workbook
import glob

# load images
for filepath in glob.iglob('jpg/*.JPG', recursive=True):
            base=os.path.basename(filepath)
            photo_number = os.path.splitext(base)[0]
            original = cv2.imread(filepath)
            mask = cv2.imread("masks/" + str(photo_number) + ".png")
            number = np.load("npy/" + str(photo_number) + ".npy")

            #resize the image
            width, height = 750,500
            resized_original = cv2.resize(original,(width,height))
            resized_mask = cv2.resize(mask,(width,height))

            n = number.shape[-1]
            for i in range (n):
                # find masks based on color
                i=i+1
                light_pink = (255, i, 255)
                dark_pink = (255, i, 255)
                mask2 = cv2.inRange(resized_mask, light_pink, dark_pink)
                result = cv2.bitwise_and(resized_mask, resized_mask, mask=mask2)

                ## find the cluster size
                na = np.array(result)
                # Find X,Y coordinates of all pink pixels
                pinkY, pinkX = np.where(np.all(na == [255, i, 255], axis=2))
                # Find first and last row containing pink pixels
                top, bottom = pinkY.min(), pinkY.max()
                # Find first and last column containing pink pixels
                left, right = pinkX.min(), pinkX.max()
                cropped=result[top:bottom,left:right]
                h,w,_ = cropped.shape

                # threshold
                gray = cv2.cvtColor(result, cv2.COLOR_BGR2GRAY)
                ret, result = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY)
                # mask on original image
                new = cv2.bitwise_and(resized_original, resized_original, mask=result)
                # gaussian
                '''gray = cv2.GaussianBlur(gray,(5,5),0)'''

                # Canny parameters
                max = 191
                min = 1
                im_canny = cv2.Canny(new, 8, 150)

                # Hough transform- circle
                circles = cv2.HoughCircles(image=im_canny, method=cv2.HOUGH_GRADIENT, dp=1, minDist=8,
                                           param1=150, param2=8, minRadius=8, maxRadius=15)
                circles = np.uint16(np.around(circles))



                #write into excel
                workbook = load_workbook("results.xlsx")
                sheets = workbook.sheetnames
                for j in sheets:
                    if j == os.path.splitext(base)[0]:
                        workbook[j].cell(row=i + 1, column=3).value = str(circles.shape[1])
                        workbook.save("results.xlsx")
                        workbook[j].cell(row=i + 1, column=7).value = h*w
                        workbook.save("results.xlsx")
                    else:
                        continue